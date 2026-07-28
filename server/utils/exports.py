"""
Shared export utilities for Django admin.

Provides helpers to generate Excel (.xlsx) and PDF responses
from a queryset + list of (header, accessor) tuples.

Usage:
    from utils.exports import export_to_excel, export_to_pdf

    def export_as_excel(modeladmin, request, queryset):
        fields = [
            ('Full Name', lambda obj: obj.full_name or ''),
            ('Email',     lambda obj: obj.email),
        ]
        return export_to_excel(queryset, fields, filename='students')
"""

import io
from datetime import datetime

from django.http import HttpResponse


# ─────────────────────────────────────────
# Excel export (openpyxl)
# ─────────────────────────────────────────

def export_to_excel(queryset, fields, filename='export'):
    """
    Build an Excel (.xlsx) HTTP response.

    :param queryset: Django QuerySet to export.
    :param fields:   List of (header_label, value_fn) tuples.
                     value_fn receives one model instance and returns a value.
    :param filename: Base filename (without extension).
    :return:         HttpResponse with attachment.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.replace('_', ' ').title()

    # ── Header styling ────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border  = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = [label for label, _ in fields]
    ws.append(headers)

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    data_align = Alignment(vertical='center', wrap_text=True)
    alt_fill   = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    for row_num, obj in enumerate(queryset, start=2):
        row_data = []
        for _, value_fn in fields:
            try:
                val = value_fn(obj)
            except Exception:
                val = ''
            # Convert non-primitives to string so Excel doesn't complain
            if hasattr(val, 'isoformat'):
                val = val.strftime('%Y-%m-%d %H:%M') if hasattr(val, 'hour') else val.isoformat()
            row_data.append(val)
        ws.append(row_data)

        # Alternate row shading
        for col_num in range(1, len(fields) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = data_align
            cell.border    = thin_border
            if row_num % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[row_num].height = 18

    # ── Auto-fit column widths ────────────────────────────────────────────────
    for col_num in range(1, len(fields) + 1):
        col_letter = get_column_letter(col_num)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_num).value or ''))
             for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # ── Build response ────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = f"{filename}_{timestamp}.xlsx"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response


# ─────────────────────────────────────────
# PDF export (reportlab)
# ─────────────────────────────────────────

def export_to_pdf(queryset, fields, filename='export', title='Export'):
    """
    Build a PDF HTTP response using reportlab.

    :param queryset: Django QuerySet to export.
    :param fields:   List of (header_label, value_fn) tuples.
    :param filename: Base filename (without extension).
    :param title:    Title printed at the top of the PDF.
    :return:         HttpResponse with attachment.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Run: pip install reportlab")

    buffer = io.BytesIO()

    # Use landscape A4 for wide tables
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1E40AF'),
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=14,
    )

    # Cell styles — used to wrap text inside table cells properly
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8,
        leading=11,          # line height
        wordWrap='CJK',      # aggressive word-wrap so long words break too
        leftPadding=0,
        rightPadding=0,
    )
    header_cell_style = ParagraphStyle(
        'HeaderCellStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.white,
        fontName='Helvetica-Bold',
        wordWrap='CJK',
        leftPadding=0,
        rightPadding=0,
    )

    elements = []

    # ── Title block ───────────────────────────────────────────────────────────
    elements.append(Paragraph(title, title_style))
    timestamp = datetime.now().strftime('%d %B %Y, %H:%M')
    elements.append(Paragraph(f"Exported on {timestamp}  |  {queryset.count()} records", subtitle_style))
    elements.append(Spacer(1, 0.3 * cm))

    # ── Build table data (Paragraph objects ensure text wraps in cells) ───────
    headers = [Paragraph(label, header_cell_style) for label in [h for h, _ in fields]]
    table_data = [headers]

    for obj in queryset:
        row = []
        for _, value_fn in fields:
            try:
                val = value_fn(obj)
            except Exception:
                val = ''
            if hasattr(val, 'strftime'):
                val = val.strftime('%Y-%m-%d %H:%M')
            # Wrap every cell in a Paragraph so reportlab handles long text
            row.append(Paragraph(str(val) if val is not None else '', cell_style))
        table_data.append(row)

    # ── Table styling ─────────────────────────────────────────────────────────
    col_count   = len(fields)
    page_width  = landscape(A4)[0] - 3 * cm           # usable width
    col_width   = page_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        # Header background
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1E40AF')),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  10),
        ('TOPPADDING',    (0, 0), (-1, 0),  10),
        # Data rows
        ('TOPPADDING',    (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),   # TOP so multi-line cells align at top
        # Alternating row backgrounds
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EFF6FF')]),
        # Grid
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('LINEBELOW',     (0, 0), (-1, 0),  1.5, colors.HexColor('#1E3A8A')),
        # Cell padding (left/right)
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = f"{filename}_{timestamp_str}.pdf"

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response
